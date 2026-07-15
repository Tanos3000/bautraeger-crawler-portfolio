# exporter.py
# Zuständig für den Excel-Export.
# Liest die gespeicherten Daten über database.py (hole_projekte/hole_einheiten/
# hole_aenderungen) und schreibt sie in eine vorformatierte Excel-Datei
# (data/wettbewerb_output.xlsx). Verwendet openpyxl für Formatierung, Farben
# und Spaltenbreiten. Dieselben Abfrage-Funktionen nutzt auch die Web-UI.

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import database

KOPF_FUELLUNG = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
KOPF_SCHRIFT = Font(bold=True, color="FFFFFF")

FARBE_VERFUEGBAR = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FARBE_VERKAUFT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FARBE_UNBEKANNT = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _formatiere_kopfzeile(blatt, spalten):
    for index, titel in enumerate(spalten, start=1):
        zelle = blatt.cell(row=1, column=index, value=titel)
        zelle.font = KOPF_SCHRIFT
        zelle.fill = KOPF_FUELLUNG
        zelle.alignment = Alignment(horizontal="center")
    blatt.freeze_panes = "A2"


def _passe_spaltenbreite_an(blatt):
    for spalte in blatt.columns:
        laengste = 0
        buchstabe = get_column_letter(spalte[0].column)
        for zelle in spalte:
            if zelle.value is not None:
                laengste = max(laengste, len(str(zelle.value)))
        blatt.column_dimensions[buchstabe].width = min(laengste + 2, 60)


def _status_farbe(status):
    if not status:
        return FARBE_UNBEKANNT
    status_klein = status.lower()
    if "verkauft" in status_klein or "vermietet" in status_klein:
        return FARBE_VERKAUFT
    if "verfügbar" in status_klein or "verkauf" in status_klein or "neubau" in status_klein:
        return FARBE_VERFUEGBAR
    return FARBE_UNBEKANNT


def _schreibe_uebersicht(arbeitsmappe):
    blatt = arbeitsmappe.active
    blatt.title = "Übersicht"
    spalten = ["Bauträger", "Projekt", "Adresse", "Einheiten", "Preis von", "Preis bis", "Status"]
    _formatiere_kopfzeile(blatt, spalten)

    for zeilen_index, projekt in enumerate(database.hole_projekte(), start=2):
        werte = [
            projekt["bautraeger"], projekt["projekt"], projekt["adresse"],
            projekt["einheiten"], projekt["preis_von"], projekt["preis_bis"], projekt["status"],
        ]
        for spalten_index, wert in enumerate(werte, start=1):
            blatt.cell(row=zeilen_index, column=spalten_index, value=wert)
        blatt.cell(row=zeilen_index, column=7).fill = _status_farbe(projekt["status"])

    _passe_spaltenbreite_an(blatt)


def _schreibe_einheiten(arbeitsmappe):
    blatt = arbeitsmappe.create_sheet("Einheiten")
    spalten = ["Projekt", "Zimmer", "Fläche m²", "Preis", "Preis/m²", "Status"]
    _formatiere_kopfzeile(blatt, spalten)

    for zeilen_index, einheit in enumerate(database.hole_einheiten(), start=2):
        werte = [
            einheit["projekt"], einheit["zimmer"], einheit["flaeche_m2"],
            einheit["preis"], einheit["preis_pro_m2"], einheit["status"],
        ]
        for spalten_index, wert in enumerate(werte, start=1):
            blatt.cell(row=zeilen_index, column=spalten_index, value=wert)
        blatt.cell(row=zeilen_index, column=6).fill = _status_farbe(einheit["status"])

    _passe_spaltenbreite_an(blatt)


def _schreibe_aenderungen(arbeitsmappe):
    blatt = arbeitsmappe.create_sheet("Änderungen")
    spalten = ["Datum", "Projekt", "Feld", "Alter Wert", "Neuer Wert"]
    _formatiere_kopfzeile(blatt, spalten)

    for zeilen_index, aenderung in enumerate(database.hole_aenderungen(), start=2):
        werte = [
            aenderung["datum"], aenderung["projekt"], aenderung["feld"],
            aenderung["alter_wert"], aenderung["neuer_wert"],
        ]
        for spalten_index, wert in enumerate(werte, start=1):
            blatt.cell(row=zeilen_index, column=spalten_index, value=wert)

    _passe_spaltenbreite_an(blatt)


def erstelle_excel():
    arbeitsmappe = Workbook()

    _schreibe_uebersicht(arbeitsmappe)
    _schreibe_einheiten(arbeitsmappe)
    _schreibe_aenderungen(arbeitsmappe)

    arbeitsmappe.save(config.PFAD_EXCEL)
    return config.PFAD_EXCEL


if __name__ == "__main__":
    pfad = erstelle_excel()
    print(f"Excel-Datei erstellt: {pfad}")
